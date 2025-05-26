import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelDataScraper:
    def __init__(self, directory=None):
        self.directory = directory
        self.results = []
        self.headers = []

    def set_directory(self, directory):
        self.directory = directory
        
    def scrape_excel_files(self, range_start="G2", range_end="G2", read_headers=True):
        """
        Scrape values from Excel files in the specified directory.
        Default is now G2 cell specifically.
        """
        if not self.directory or not os.path.isdir(self.directory):
            raise ValueError("Please set a valid directory first.")
            
        self.results = []
        self.headers = []
        
        # Extract column letters from range
        start_col = range_start[0]
        end_col = range_end[0]
        
        # Extract row numbers from range
        start_row = int(range_start[1:])
        end_row = int(range_end[1:])
        
        # Convert column letters to indices
        start_col_idx = ord(start_col) - ord('A') + 1
        end_col_idx = ord(end_col) - ord('A') + 1
        
        for filename in os.listdir(self.directory):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(self.directory, filename)
                try:
                    workbook = load_workbook(file_path, data_only=True)
                    sheet = workbook.active
                    
                    # Read headers if specified
                    if read_headers and not self.headers:
                        header_row = []
                        for col_idx in range(start_col_idx, end_col_idx + 1):
                            col_letter = get_column_letter(col_idx)
                            cell_value = sheet[f"{col_letter}1"].value
                            header_row.append(cell_value if cell_value else f"Column {col_letter}")
                        self.headers = header_row
                    
                    # Read data from specified range
                    file_data = {"filename": filename, "values": {}}
                    for row in range(start_row, end_row + 1):
                        row_data = []
                        for col_idx in range(start_col_idx, end_col_idx + 1):
                            col_letter = get_column_letter(col_idx)
                            cell_value = sheet[f"{col_letter}{row}"].value
                            row_data.append(cell_value)
                            
                        # Map data to headers if available
                        if self.headers:
                            for idx, header in enumerate(self.headers):
                                if idx < len(row_data):
                                    file_data["values"][header] = row_data[idx]
                        else:
                            # Use column letters as keys if no headers
                            for idx, value in enumerate(row_data):
                                col_letter = get_column_letter(start_col_idx + idx)
                                file_data["values"][col_letter] = value
                    
                    self.results.append(file_data)
                except Exception as e:
                    print(f"Error processing {filename}: {str(e)}")
                    
        return self.results
    
    def get_headers(self):
        return self.headers
    
    def get_results(self):
        return self.results
    
    def save_results_to_csv(self, output_file):
        """Save scraped results to a CSV file"""
        if not self.results:
            print("No results to save.")
            return False
            
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                # Determine headers
                if self.headers:
                    fieldnames = ['filename'] + self.headers
                else:
                    # Use the keys from the first result's values
                    fieldnames = ['filename'] + list(self.results[0]['values'].keys())
                
                import csv
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in self.results:
                    row = {'filename': result['filename']}
                    row.update(result['values'])
                    writer.writerow(row)
                    
            print(f"Results saved to {output_file}")
            return True
        except Exception as e:
            print(f"Error saving results: {str(e)}")
            return False
        
    def save_results_to_excel(self, output_file):
        """Save scraped results to an Excel file"""
        if not self.results:
            print("No results to save.")
            return False
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Combined Data"
            
            # Write headers
            if self.headers:
                headers = ['filename'] + self.headers
            else:
                # Use the keys from the first result's values
                headers = ['filename'] + list(self.results[0]['values'].keys())
                
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, result in enumerate(self.results, 2):
                ws.cell(row=row_idx, column=1, value=result['filename'])
                
                for col_idx, header in enumerate(headers[1:], 2):
                    value = result['values'].get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
                    
            wb.save(output_file)
            print(f"Results saved to Excel file: {output_file}")
            return True
        except Exception as e:
            print(f"Error saving results to Excel: {str(e)}")
            return False