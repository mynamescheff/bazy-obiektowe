from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class ExcelTransposer:
    def __init__(self, filename: str):
        self.filename = filename
        self.workbook: Workbook = load_workbook(filename)
        self.sheet: Worksheet = self.workbook.active

    def set_active_sheet(self, sheet_name: str) -> None:
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    def transpose_cells_to_table(self) -> None:
        try:
            data = []
            for row in self.sheet.iter_rows(values_only=True):
                data.append(row)
            transposed_data = list(map(list, zip(*data)))
            transposed_sheet = self.workbook.create_sheet(title="Transposed")
            
            for row_idx, row_data in enumerate(transposed_data):
                for col_idx, cell_value in enumerate(row_data):
                    column_letter = get_column_letter(col_idx + 1)
                    transposed_sheet[f"{column_letter}{row_idx + 1}"] = cell_value
                    
            self.auto_adjust_column_width(transposed_sheet)
            self.workbook.save(self.filename)
        except Exception as e:
            print(f"An error occurred while transposing the data: {e}")

    def auto_adjust_column_width(self, sheet: Worksheet) -> None:
        for column in sheet.columns:
            max_length = 0
            column_cells = [cell for cell in column]
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width