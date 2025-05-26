import os
import re
import time
import shutil
import datetime
from datetime import date, datetime
from pathlib import Path
from tkinter import Tk, Frame, Label, Button, Entry, Text, Toplevel, filedialog, messagebox, Scrollbar, RIGHT, Y, \
    Checkbutton, BooleanVar, W, E
from tkinter import *
import tkinter as tk
import win32com.client
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import queue
import threading
import sys
import pythoncom
import csv
import openpyxl
from openpyxl.cell.cell import MergedCell

SHARED_MAILBOX_EMAIL = ""
notes_window = None
instructions_window = None
notes_content = ""
notes_first_time = True


class CharacterTransformer:
    def __init__(self):
        self.character_mapping = {
            'á': 'a', 'à': 'a', 'â': 'a', 'ä': 'a', 'ã': 'a', 'å': 'a', 'æ': 'ae', 'ç': 'c', 'č': 'c', 'ć': 'c',
            'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e', 'ė': 'e', 'í': 'i', 'ì': 'i', 'î': 'i', 'ï': 'i', 'ñ': 'n',
            'ó': 'o', 'ò': 'o', 'ô': 'o', 'ö': 'o', 'õ': 'o', 'ø': 'o', 'œ': 'oe', 'š': 's', 'ß': 'ss', 'ú': 'u',
            'ù': 'u', 'û': 'u', 'ü': 'u', 'ý': 'y', 'ÿ': 'y', 'ž': 'z', 'Á': 'A', 'À': 'A', 'Â': 'A', 'Ä': 'A',
            'Ã': 'A', 'Å': 'A', 'Æ': 'AE', 'Ç': 'C', 'Č': 'C', 'Ć': 'C', 'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
            'Ė': 'E', 'Í': 'I', 'Ì': 'I', 'Î': 'I', 'Ï': 'I', 'Ñ': 'N', 'Ó': 'O', 'Ò': 'O', 'Ô': 'O', 'Ö': 'O',
            'Õ': 'O', 'Ø': 'O', 'Œ': 'OE', 'Š': 'S', 'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U', 'Ý': 'Y', 'Ÿ': 'Y',
            'Ž': 'Z', 'Þ': 'th', 'þ': 'th', 'ð': 'dh', 'Đ': 'D', 'đ': 'd', 'ł': 'l', 'Ł': 'L', 'ū': 'u', 'Ū': 'U',
            'Ā': 'A', 'ā': 'a', 'Ē': 'E', 'ē': 'e', 'Ī': 'I', 'ī': 'i', 'Ō': 'O', 'ō': 'o', 'ă': 'a', 'Ă': 'A',
            'ș': 's', 'Ș': 'S', 'ț': 't', 'Ț': 'T', 'â': 'a', 'Â': 'A', 'î': 'i', 'Î': 'I', 'ş': 's', 'Ş': 'S',
            'ğ': 'g', 'Ğ': 'G', 'İ': 'I', 'ı': 'i', 'ö': 'o', 'Ö': 'O', 'ü': 'u', 'Ü': 'U', 'ț': 't', 'Ţ': 'T',
            'ţ': 't',

            '_': ' ', '"': ' ', '  ': ' ', '\xa0': ' ', '\t': ' ', '\n': ' ', '\r': ' ', '\x0b': ' ', '\x0c': ' ',
            '\u200b': ' ', '\u200c': ' ', '\u200d': ' ', '\u200e': ' ', '\u200f': ' ', '\u202a': ' ', '\u202c': ' ',
            '\u202d': ' ', '\u202e': ' ', '\u202f': ' ', '\u205f': ' ', '\u3000': ' ', '\u2000': ' ', '\u2001': ' ',
            '\u2002': ' ', '\u2003': ' ', '\u2004': ' ', '\u2005': ' ', '\u2006': ' ', '\u2007': ' ', '\u2008': ' ',
            '\u2009': ' ', '\u200a': ' ', ' ': ' ', ' ': ' ', '&nbsp;': '', '\u0219': 's', 'ș': 's', 'ă': 'a', 'ț': 't',
            'ő': 'o', 'Ő': 'O', 'ű': 'u', 'Ű': 'U', '\n': ' ', '\r': ' ', '\t': ' ', '\xa0': ' ', '\u200b': ' ',
            '\u200c': ' ',
            'ę': 'e', 'Ę': 'E', 'ą': 'a', 'Ą': 'A', 'ś': 's', 'Ś': 'S', 'ł': 'l', 'Ł': 'L', 'ż': 'z', 'Ż': 'Z',
            'ź': 'z', 'Ź': 'Z',
            ';': '', 'ń': 'n',
        }

    def transform_to_swift_accepted_characters(self, input_list):
        transformed_list = []
        for input_string in input_list:
            transformed_string = re.sub(r'\b\w+\b',
                                        lambda m: ''.join(self.character_mapping.get(char, char) for char in m.group()),
                                        str(input_string))
            transformed_string = re.sub(r'[.,]', '', transformed_string)
            transformed_list.append(transformed_string)
        return transformed_list


class Wide:
    def __init__(self, file_path: str, sheet_name: str, directory: str):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.directory = directory
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def auto_adjust_column_width(self) -> None:
        for column in self.sheet.columns:
            max_length = 0
            column_cells = [cell for cell in column]
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
        self.workbook.save(self.file_path)
        print("Column widths adjusted and saved successfully.")

    def get_file_count(self) -> int:
        file_count = len(
            [name for name in os.listdir(self.directory) if os.path.isfile(os.path.join(self.directory, name))])
        return file_count

    def create_table_with_headers(self, table_name: str) -> None:
        file_count = self.get_file_count()
        if file_count == 0:
            print("No files found in the directory.")
            return
        last_row = file_count + 1
        last_column = self.sheet.max_column
        last_column_letter = get_column_letter(last_column)
        new_sheet = self.workbook.create_sheet(title=f"{table_name}_Sheet")
        for row in self.sheet.iter_rows(max_row=last_row, max_col=last_column):
            new_sheet.append([cell.value for cell in row])
        self.auto_adjust_column_width(new_sheet)
        table_range = f"A1:{last_column_letter}{last_row}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        new_sheet.add_table(table)
        self.workbook.save(self.file_path)
        print("Table with headers created in new sheet and saved successfully.")


class ExcelComparator:
    def __init__(self, combined_file, sprawdzacz_file, output_path):
        self.combined_file = combined_file
        self.sprawdzacz_file = sprawdzacz_file
        self.output_path = output_path

    def compare_and_append(self):
        try:
            combined_workbook = load_workbook(filename=self.combined_file)
            combined_sheet = combined_workbook['Transposed']
            sprawdzacz_workbook = load_workbook(filename=self.sprawdzacz_file)
            sprawdzacz_sheet = sprawdzacz_workbook['name_acc']
            non_matching_values = []
            filenames = []

            # Check and rename duplicates in column B
            b_values = {}
            for row in range(2, combined_sheet.max_row + 1):
                b_value = combined_sheet[f'B{row}'].value

                if b_value in b_values:
                    b_values[b_value].append(row)
                else:
                    b_values[b_value] = [row]

            for b_value, rows in b_values.items():
                if len(rows) > 1:
                    counter = 2  # Start numbering from 2
                    for row in rows:
                        a_value = combined_sheet[f'A{row}'].value.replace(" ", "")
                        if any(char.isdigit() for char in a_value):
                            combined_sheet[f'B{row}'].value = f"{b_value}{counter}"
                            counter += 1

            # Save the workbook after renaming
            combined_workbook.save(self.combined_file)

            # Compare values as before
            for row in range(2, combined_sheet.max_row + 1):
                case_value = combined_sheet[f'A{row}'].value
                f_value = str(combined_sheet[f'F{row}'].value).replace(" ", "").replace("-", "") if combined_sheet[
                    f'F{row}'].value else ""
                h_value = str(combined_sheet[f'G{row}'].value).replace(" ", "").replace("-", "") if combined_sheet[
                    f'G{row}'].value else ""
                found_match = any(f_value == str(row_val[2]) or h_value == str(row_val[2]) for row_val in
                                  sprawdzacz_sheet.iter_rows(values_only=True))
                if not found_match:
                    non_matching_values.append((case_value, f_value, h_value))
                    filenames.append((self.combined_file, self.sprawdzacz_file))

            file_path = os.path.join(self.output_path, 'utils\\mismatch_list.txt')
            with open(file_path, 'w', encoding="utf-8") as file:
                current_date = date.today()
                for value in non_matching_values:
                    modified_value = value[0].replace("\xa0", "").strip()
                    file.write(f"{modified_value}: {value[1]}, {value[2]} ({current_date})\n")

            return non_matching_values, f"Non-matching values appended to {file_path} successfully."
        except Exception as e:
            return [], str(e)


class CaseList:
    def __init__(self, excel_folder, list_folder):
        self.excel_folder = excel_folder
        self.list_folder = list_folder

    def process_excel_files(self):
        list_file_path = os.path.join(self.list_folder, "case_list.txt")
        existing_values = {}
        duplicate_counts = {}
        error_messages = []
        if os.path.isfile(list_file_path):
            existing_values, duplicate_counts = self.load_existing_list(list_file_path)
        all_entries = []
        for file_name in os.listdir(self.excel_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(self.excel_folder, file_name)
                try:
                    wb = load_workbook(file_path)
                    sheet = wb.active
                    value = sheet["B17"].value
                    if value in existing_values:
                        value.replace(" ", "").replace("\n", "").replace("\t", "").replace("\r", "").replace("  ",
                                                                                                             " ").replace(
                            " ", "").replace("\xa0", "").replace(" ", "").lstrip().replace("\n", " ").replace("\t",
                                                                                                              " ").replace(
                            "\r", " ").replace("\x0b", " ").replace("\x0c", " ").replace("\t", " ").replace("\r",
                                                                                                            " ").replace(
                            "\x85", " ").replace("\r\n", " ").replace("\u2028", " ").replace("\u2029", " ")
                        value.replace(" ", "")
                        # make sure that the string of the value variable ends on a number or a letter
                        while not value[-1].isalnum():
                            value = value[:-1]
                        value = re.sub(r'[\n\r\t\v\f\x85\u2028\u2029]+', ' ', value)
                        # keep replacing double spaces from the string until there are only singular space overall
                        while "  " in value:
                            value = value.replace("  ", " ")
                        while "\n" in value:
                            value = value.replace("\n", "")
                        duplicate_counts[value] += 1
                        existing_values[value] = True
                        entry = f"{value} [{file_name} - DUPLICATE {duplicate_counts[value]}]"
                        all_entries.append(entry)
                        print(
                            f"Duplicate found: {value} in file {file_name} (Duplicate count: {duplicate_counts[value]})")
                    else:
                        existing_values[value] = False
                        duplicate_counts[value] = 0
                        entry = f"{value} [{file_name}]"
                        all_entries.append(entry)
                except Exception as e:
                    error_messages.append(f"Error processing file '{file_name}': {str(e)}")
        if all_entries:
            today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(list_file_path, "a", encoding="utf-8") as file:
                file.write(f"\n--- Updated on {today} ---\n")
                for entry in all_entries:
                    file.write(f"{entry} ({today})\n")
        print("Cases processed successfully and saved to the case_list.txt file.")
        return duplicate_counts, error_messages

    def load_existing_list(self, list_file_path):
        existing_values = {}
        duplicate_counts = {}
        with open(list_file_path, "r", encoding="utf-8") as file:
            for line in file:
                if line.strip() and not line.startswith("---"):
                    parts = line.split(" [")
                    value = parts[0]
                    existing_values[value] = False
                    duplicate_count = parts[1].count("DUPLICATE")
                    duplicate_counts[value] = duplicate_count
        return existing_values, duplicate_counts


def check_file_conditions(excel_file_name, cell_b20_value, cell_c20_value):
    is_condition_1 = cell_b20_value == 18 and cell_c20_value == "GBP"
    is_condition_2 = "name1" in excel_file_name.upper() and cell_b20_value == 15
    is_condition_3 = "name2" in excel_file_name.upper() and cell_b20_value == 15
    is_condition_4 = "name3" in excel_file_name.upper() and cell_b20_value == 15
    if is_condition_1 or is_condition_2 or is_condition_3 or is_condition_4:
        return True, None
    else:
        mismatched_values = {"cell_b20_value": cell_b20_value, "cell_c20_value": cell_c20_value}
        return False, mismatched_values


class OutlookProcessor:
    def __init__(self, category, target_senders, attachment_save_path, msg_save_path):
        self.category = category
        self.target_senders = target_senders
        self.attachment_save_path = attachment_save_path
        self.msg_save_path = msg_save_path
        self.outlook = win32com.client.Dispatch("Outlook.Application")
        self.namespace = self.outlook.GetNamespace("MAPI")
        os.makedirs(self.attachment_save_path, exist_ok=True)
        os.makedirs(self.msg_save_path, exist_ok=True)

    def list_unread_emails(self):
        recipient = self.namespace.CreateRecipient(SHARED_MAILBOX_EMAIL)
        recipient.Resolve()
        if recipient.Resolved:
            shared_mailbox = self.namespace.GetSharedDefaultFolder(recipient, 6)
            unread_emails = shared_mailbox.Items.Restrict(f"[Categories] = '{self.category}' AND [UnRead] = True")
            return len([email for email in unread_emails])
        return 0

    def get_unique_filename(self, base_path, original_filename, extension):
        counter = 2
        new_filename = original_filename
        while os.path.exists(os.path.join(base_path, f"{new_filename}{extension}")):
            new_filename = f"{original_filename} {counter}"
            counter += 1
        return new_filename

    def mark_email_as_read(self, item, mark_as_read):
        item.UnRead = not mark_as_read
        item.Save()
        if mark_as_read:
            print("Marked email as read.")
        else:
            print("Email left as unread.")

    def download_attachments_and_save_as_msg(self, save_emails, mark_as_read):
        recipient = self.namespace.CreateRecipient(SHARED_MAILBOX_EMAIL)
        recipient.Resolve()
        self.processed_emails = {}  # Clear previous tracking
        self.emails_with_pdf = []  # Clear previous tracking
        self.emails_with_nvf_new_vendor = []  # Clear previous tracking
        emails_with_no_attachments = []  # Track emails with no attachments
        if recipient.Resolved:
            shared_mailbox = self.namespace.GetSharedDefaultFolder(recipient, 6)
            unread_emails = shared_mailbox.Items.Restrict(f"[Categories] = '{self.category}' AND [UnRead] = True")
            emails_to_process = [email for email in unread_emails]
            print(
                f"Found {len(emails_to_process)} emails to process under category '{self.category}' with 'UnRead' = True.")
            if save_emails:
                saved_emails = 0
                saved_attachments = 0
                not_saved_subjects = []
                incorrect_subjects = []
                for item in emails_to_process:
                    time.sleep(2)
                    try:
                        sender_email = item.SenderEmailAddress
                        sender_name_match = re.search(
                            r'/O=EXCHANGELABS/OU=EXCHANGE ADMINISTRATIVE GROUP.*?-([A-Za-z]+)', sender_email)
                        sender_name = sender_name_match.group(1) if sender_name_match else sender_email
                        print(f"Processing email from: {sender_name}")
                        if sender_name.lower() in [sender.lower() for sender in self.target_senders]:
                            if item.Attachments.Count == 0:
                                # Email has no attachments
                                emails_with_no_attachments.append(item.Subject)
                                print(
                                    f"Email with subject '{item.Subject}' has no attachments and will not be processed.")
                                continue  # Skip to the next email

                            # If we reach here, the email has attachments
                            subject_correct = True
                            saved_attachment_paths = []
                            has_pdf = False
                            has_nvf_new_vendor = False
                            save_msg_once = False
                            if item.Attachments.Count > 0:
                                for attachment in item.Attachments:
                                    if attachment.FileName.lower().endswith('.pdf'):
                                        has_pdf = True
                                    if 'nvf' in attachment.FileName.lower() or 'new vendor' in attachment.FileName.lower() or 'vendor' in attachment.FileName.lower():
                                        has_nvf_new_vendor = True
                                    if attachment.FileName.lower().endswith('.xlsx'):
                                        new_filename = self.extract_filename_from_subject(item.Subject)
                                        transformer = CharacterTransformer()
                                        new_filename = \
                                        transformer.transform_to_swift_accepted_characters([new_filename])[0]
                                        new_filename = re.sub(r'[\/:*?"<>|\t]', ' ', new_filename)
                                        new_filename = re.sub(r'[^A-Za-z0-9\s\-\–;]', '', new_filename)
                                        if ';' not in new_filename and ';' not in item.Subject:
                                            print(f"Invalid filename generated from subject: {new_filename}")
                                            subject_correct = False
                                            incorrect_subjects.append(item.Subject)
                                            continue
                                        unique_attachment_filename = self.get_unique_filename(self.attachment_save_path,
                                                                                              new_filename, '.xlsx')
                                        attachment_path = os.path.join(self.attachment_save_path,
                                                                       f"{unique_attachment_filename}.xlsx")
                                        attachment.SaveAsFile(attachment_path)
                                        print(f"Saved attachment: {attachment_path}")
                                        saved_attachment_paths.append(attachment_path)  # Track saved attachments
                                        saved_attachments += 1
                                        if not save_msg_once:
                                            unique_msg_filename = self.get_unique_filename(self.msg_save_path,
                                                                                           new_filename,
                                                                                           ' approval.msg')
                                            approval_msg_path = os.path.join(self.msg_save_path,
                                                                             f"{unique_msg_filename} approval.msg")
                                            item.SaveAs(approval_msg_path)
                                            print(f"Saved Outlook message: {approval_msg_path}")
                                            save_msg_once = True  # Ensure only one message file is saved
                                if has_pdf:
                                    self.emails_with_pdf.append(item.Subject)
                                if has_nvf_new_vendor:
                                    self.emails_with_nvf_new_vendor.append(item.Subject)
                                    print(f"NVF is present in the email with subject: {item.Subject}")
                                if subject_correct and not has_nvf_new_vendor:
                                    self.mark_email_as_read(item, mark_as_read)
                                    saved_emails += 1
                                    self.processed_emails[
                                        item.Subject] = saved_attachment_paths  # Track processed email
                                elif has_nvf_new_vendor:
                                    self.clean_up_files(saved_attachment_paths, approval_msg_path)
                    except Exception as e:
                        print(f"Error processing email from {sender_email}: {str(e)}")
                        not_saved_subjects.append(item.Subject)
                print(f"Total saved emails: {saved_emails}")
                actual_saved_attachments = self.count_files_in_directory(self.attachment_save_path)
                print(f"Total saved attachments: {actual_saved_attachments}")
                if not_saved_subjects:
                    print("Emails not saved due to errors:")
                    for subject in not_saved_subjects:
                        print(subject)
                if incorrect_subjects:
                    print("Emails with incorrect subject format:")
                    for subject in incorrect_subjects:
                        print(subject)
                # Additional block to copy attachments to the utils/pmt_run directory
                self.copy_attachments_to_pmt_run()

                # Create log file
                log_directory = r"C:\IT project3\utils\logs"
                os.makedirs(log_directory, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                log_filename = f"process_log_{timestamp}.txt"
                log_path = os.path.join(log_directory, log_filename)

                # Instantiate CharacterTransformer
                transformer = CharacterTransformer()

                with open(log_path, 'w', encoding='utf-8') as log_file:
                    log_file.write("Summary of processed emails and saved attachments:\n")
                    if self.processed_emails:
                        for subject, attachments in self.processed_emails.items():
                            # Transform the subject
                            transformed_subject = transformer.transform_to_swift_accepted_characters([subject])[0]
                            log_file.write(f"Email subject: {transformed_subject}\n")
                            # Transform the attachment names
                            transformed_attachments = transformer.transform_to_swift_accepted_characters(attachments)
                            for attachment in transformed_attachments:
                                log_file.write(f" - Saved attachment: {attachment}\n")

                    actual_saved_attachments = self.count_files_in_directory(self.attachment_save_path)
                    if saved_attachments != actual_saved_attachments:
                        log_file.write(
                            f"Discrepancy detected: expected {saved_attachments} attachments, but found {actual_saved_attachments} in directory.\n")

                    if self.emails_with_pdf:
                        log_file.write("Emails with PDF attachments:\n")
                        for subject in self.emails_with_pdf:
                            log_file.write(f"{subject}\n")
                        print("PDF attachments are present in the following emails:")
                        for subject in self.emails_with_pdf:
                            print(subject)

                    if self.emails_with_nvf_new_vendor:
                        log_file.write("Emails with NVF or New Vendor in attachment filenames:\n")
                        for subject in self.emails_with_nvf_new_vendor:
                            log_file.write(f"{subject}\n")
                        print("NVF or New Vendor is present in the following emails:")
                        for subject in self.emails_with_nvf_new_vendor:
                            print(subject)

                    if emails_with_no_attachments:
                        log_file.write("Emails with no attachments:\n")
                        for subject in emails_with_no_attachments:
                            log_file.write(f"{subject}\n")
                        print("Emails with no attachments:")
                        for subject in emails_with_no_attachments:
                            print(subject)

                    if not_saved_subjects:
                        print("Emails not saved due to errors:")
                        for subject in not_saved_subjects:
                            print(subject)

                print(f"Log file created at {log_path}")

            else:
                print("Email saving is disabled. No emails were saved.")
        else:
            print(f"Could not resolve the recipient: {SHARED_MAILBOX_EMAIL}")

    def clean_up_files(self, attachment_paths, msg_path):
        for attachment_path in attachment_paths:
            try:
                os.remove(attachment_path)
                print(f"Deleted attachment: {attachment_path}")
            except Exception as e:
                print(f"Error deleting attachment {attachment_path}: {str(e)}")
        try:
            os.remove(msg_path)
            print(f"Deleted Outlook message: {msg_path}")
        except Exception as e:
            print(f"Error deleting Outlook message {msg_path}: {str(e)}")

    def copy_attachments_to_pmt_run(self):
        # Get the absolute path of the script
        script_path = os.path.dirname(os.path.abspath(__file__))
        # Define the pmt_run directory relative to the script location
        pmt_run_dir = os.path.join(script_path, "utils", "pmt_run")
        os.makedirs(pmt_run_dir, exist_ok=True)
        for filename in os.listdir(self.attachment_save_path):
            source_file = os.path.join(self.attachment_save_path, filename)
            destination_file = os.path.join(pmt_run_dir, filename)
            shutil.copy(source_file, destination_file)
        print(f"All files copied to {pmt_run_dir}")

    def extract_filename_from_subject(self, subject):
        match = re.search(r';\s*(.*)', subject)
        if match:
            return match.group(1)
        else:
            return subject

    def count_files_in_directory(self, directory):
        return len([name for name in os.listdir(directory) if os.path.isfile(os.path.join(directory, name))])


class ExcelTransposer:
    def __init__(self, filename: str):
        self.filename = filename
        self.workbook: Workbook = load_workbook(filename)
        self.sheet: Worksheet = self.workbook.active

    def set_active_sheet(self, sheet_name: str) -> None:
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    def transpose_cells_to_table(self) -> None:
        try:
            data = []
            for row in self.sheet.iter_rows(values_only=True):
                data.append(row)
            transposed_data = list(map(list, zip(*data)))
            transposed_sheet = self.workbook.create_sheet(title="Transposed")
            for row_idx, row_data in enumerate(transposed_data):
                for col_idx, cell_value in enumerate(row_data):
                    column_letter = get_column_letter(col_idx + 1)
                    transposed_sheet[f"{column_letter}{row_idx + 1}"] = cell_value
            self.auto_adjust_column_width(transposed_sheet)
            # if the values in any row from columns F and G are both "None", print a warning with naming the row from value of the A column in that row
            for row in transposed_sheet.iter_rows(min_row=2, max_row=transposed_sheet.max_row, min_col=6, max_col=7):
                if row[0].value == "None" and row[1].value == "None":
                    print(
                        f"Warning: There are missing bank account information in row {row[0].row}. The filename is {transposed_sheet[f'A{row[0].row}'].value}")
            # hide H column in the transposed sheet
            transposed_sheet.column_dimensions['H'].hidden = True
            self.workbook.save(self.filename)
            print("Transposed data saved and column widths adjusted successfully.")
        except Exception as e:
            print(f"An error occurred while transposing the data: {e}")

    def auto_adjust_column_width(self, sheet: Worksheet) -> None:
        for column in sheet.columns:
            max_length = 0
            column_cells = [cell for cell in column]
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


class KejsarProcessor:
    def __init__(self, base_dir: str):
        self.source_dir = os.path.join(base_dir, "utils", "pmt_run")
        self.error_dir = os.path.join(self.source_dir, "error_files")
        self.combined_file = os.path.join(base_dir, "utils", "combined_file.xlsx")
        self.case_list_folder = os.path.join(base_dir, "utils")
        self.comparison_file = os.path.join(base_dir, "utils", "comparison_file.xlsx")
        self.excel_files = list(Path(self.source_dir).glob("*.xlsx"))
        self.mismatched_cases = []
        os.makedirs(self.error_dir, exist_ok=True)
        self.comparison_data = self.load_comparison_data()

    def load_comparison_data(self):
        """Load data from comparison_file.xlsx with both cleaned and original values."""
        comparison_data = {}
        wb = load_workbook(filename=self.comparison_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Store cleaned key -> original bank name mapping
            if row[2] is not None:  # Column C value
                cleaned_key = str(row[2]).replace(" ", "").replace("-", "")
                comparison_data[cleaned_key] = row[0]  # Column A value (original bank name)
        return comparison_data

    def clean_value(self, value):
        """Minimal cleaning just for matching keys."""
        if value is not None:
            return str(value).replace(" ", "").replace("-", "")
        return value

    def match_bank_name(self, extra_cell_3_value, extra_cell_4_value):
        """Match cleaned values against comparison data and return original bank name."""
        if extra_cell_3_value is not None:
            extra_cell_3_value = str(extra_cell_3_value).replace(" ", "").replace("-", "")
        if extra_cell_4_value is not None:
            extra_cell_4_value = str(extra_cell_4_value).replace(" ", "").replace("-", "")

        # Look up original bank name using cleaned values
        bank_name = self.comparison_data.get(extra_cell_3_value) or self.comparison_data.get(extra_cell_4_value)
        return bank_name

    def process_files(self):
        i = 0
        while i < len(self.excel_files):
            excel_file = self.excel_files[i]
            try:
                wb = load_workbook(filename=excel_file)
                sheet_name = "Sheet1"
                if sheet_name not in wb.sheetnames:
                    print(f"Skipping and moving file {excel_file.name} as '{sheet_name}' worksheet does not exist.")
                    shutil.move(str(excel_file), os.path.join(self.error_dir, excel_file.name))
                    self.excel_files.pop(i)
                    continue
                ws = wb[sheet_name]
                cell_b20_value = ws["B20"].value
                cell_c20_value = ws["C20"].value
                is_condition_met, mismatched_values = check_file_conditions(excel_file.name, cell_b20_value,
                                                                            cell_c20_value)
                if not is_condition_met:
                    self.mismatched_cases.append((excel_file.name, mismatched_values))
            except Exception as e:
                print(f"An error occurred with file {excel_file}: {e}")
                shutil.move(str(excel_file), os.path.join(self.error_dir, excel_file.name))
                self.excel_files.pop(i)
                continue
            i += 1

    def print_mismatched_cases(self):
        if self.mismatched_cases:
            print("Cases with mismatched values in B20 and C20 cells:")
            for case in self.mismatched_cases:
                print(f"File: {case[0]}, Mismatched Values: {case[1]}")

    def collect_values(self):
        values_excel_files = {}
        for excel_file in self.excel_files:
            wb = load_workbook(filename=excel_file)
            extra_cell_1 = wb["Sheet1"]["B19"]
            extra_cell_2 = wb["Sheet1"]["C19"]
            # if AQA is only in the first 10 characters of the filename
            if "AQA" in excel_file.name[
                        :10] or "Harrow School" in excel_file.name or "Assessment Qualification Alliance" in excel_file.name:
                extra_cell_3 = wb["Sheet1"]["C24"]
                extra_cell_4 = wb["Sheet1"]["C25"]
                extra_cell_5 = wb["Sheet1"]["C26"]
            else:
                extra_cell_3 = wb["Sheet1"]["C33"]
                extra_cell_4 = wb["Sheet1"]["C34"]
                extra_cell_5 = wb["Sheet1"]["C35"]
            rng_cell_1 = wb["Sheet1"]["B16"]
            rng_cell_2 = wb["Sheet1"]["B17"]
            # vendor name
            rng_cell_3 = wb["Sheet1"]["B8"]
            # Clean up the cells, but only modify if they are not MergedCells
            if not isinstance(rng_cell_1, MergedCell):
                rng_cell_1.value = str(rng_cell_1.value).replace("  ", " ").replace("\xa0", "").replace(" ",
                                                                                                        "").replace(".",
                                                                                                                    " ").replace(
                    "–", "-").lstrip().replace("\n", " ").replace("\t", " ").replace("\r", " ").replace("\x0b",
                                                                                                        " ").replace(
                    "\x0c", " ").replace("\t", " ").replace("\r", " ").replace("\x85", " ").replace("\r\n",
                                                                                                    " ").replace(
                    "\u2028", " ").replace("\u2029", " ").replace(";", " ")

            if not isinstance(rng_cell_2, MergedCell):
                rng_cell_2.value = str(rng_cell_2.value).replace(" ", "").replace("\xa0", "").replace(" ",
                                                                                                      "").lstrip().replace(
                    "\n", " ").replace("\t", " ").replace("\r", " ").replace("\x0b", " ").replace("\x0c", " ").replace(
                    "\t", " ").replace("\r", " ").replace("\x85", " ").replace("\r\n", " ").replace("\u2028",
                                                                                                    " ").replace(
                    "\u2029", " ")

            if not isinstance(rng_cell_3, MergedCell):
                rng_cell_3.value = str(rng_cell_3.value).replace("\xa0", "").replace(" ", "").replace("\n",
                                                                                                      " ").replace("\t",
                                                                                                                   " ").replace(
                    "\r", " ").replace("\x0b", " ").replace("\x0c", " ").replace("\t", " ").replace("\r", " ").replace(
                    "\x85", " ").replace("\r\n", " ").replace("\u2028", " ").replace("\u2029", " ")

            rng_values = [rng_cell_1.value, rng_cell_2.value, rng_cell_3.value]
            # Clean up the cells, but only if they are not MergedCells

            if not isinstance(extra_cell_3, MergedCell):
                extra_cell_3.value = str(extra_cell_3.value).replace("/", "").lstrip().replace(" ",
                                                                                               "").rstrip().replace(
                    "  ", "").replace("\n", " ").replace("\t", " ").replace("\r", " ").replace("\x0b", " ").replace(
                    "\x0c", " ").replace("\t", " ").replace("\r", " ").replace("\x85", " ").replace("\r\n",
                                                                                                    " ").replace(
                    "\u2028", " ").replace("\u2029", " ").replace(" ", "").replace("IBAN", "").replace("IBAN:",
                                                                                                       "").replace(":",
                                                                                                                   "")
            if not isinstance(extra_cell_4, MergedCell):
                extra_cell_4.value = str(extra_cell_4.value).replace("/", "").lstrip().replace(" ",
                                                                                               "").rstrip().replace(
                    "  ", "").replace("\n", " ").replace("\t", " ").replace("\r", " ").replace("\x0b", " ").replace(
                    "\x0c", " ").replace("\t", " ").replace("\r", " ").replace("\x85", " ").replace("\r\n",
                                                                                                    " ").replace(
                    "\u2028", " ").replace("\u2029", " ").replace(" ", "")
            # replace characters "USD" with "" in extra_cell_3 and extra_cell_4, but only if those are the first three characters
            if extra_cell_3.value is not None:
                extra_cell_3.value = extra_cell_3.value.replace("USD", "") if extra_cell_3.value[
                                                                              :3] == "USD" else extra_cell_3.value
            if extra_cell_4.value is not None:
                extra_cell_4.value = extra_cell_4.value.replace("USD", "") if extra_cell_4.value[
                                                                              :3] == "USD" else extra_cell_4.value
            # replace characters "EUR" with "" in extra_cell_4 and extra_cell_3, but only if those are the last three characters
            if extra_cell_4.value is not None:
                extra_cell_4.value = extra_cell_4.value.replace("EUR", "") if extra_cell_4.value[
                                                                              -3:] == "EUR" else extra_cell_4.value
            if extra_cell_3.value is not None:
                extra_cell_3.value = extra_cell_3.value.replace("EUR", "") if extra_cell_3.value[
                                                                              -3:] == "EUR" else extra_cell_3.value

            if extra_cell_4.value is not None:
                extra_cell_4.value = extra_cell_4.value.replace("/", "").lstrip().replace(" ", "").rstrip().replace(
                    "  ", "").replace("\n", " ").replace("\t", " ").replace("\r", " ").replace("\x0b", " ").replace(
                    "\x0c", " ").replace("\t", " ").replace("\r", " ").replace("\x85", " ").replace("\r\n",
                                                                                                    " ").replace(
                    "\u2028", " ").replace("\u2029", " ").replace(" ", "")
            if extra_cell_2.value is not None:
                extra_cell_2.value = extra_cell_2.value.replace(" ", "").replace("\n", "").replace("\xa0", "").replace(
                    "\t", "").replace(',', '.').replace('RS', 'PKR').replace('Rs', 'PKR').replace('rs', 'PKR').replace(
                    'EURO', 'EUR')

            transformer = CharacterTransformer()
            transformed_values = transformer.transform_to_swift_accepted_characters(rng_values)

            extra_cell_1_value = extra_cell_1.value
            extra_cell_2_value = extra_cell_2.value
            extra_cell_2_value = str(extra_cell_2_value).lstrip()
            extra_cell_3_value = self.clean_value(extra_cell_3.value)
            extra_cell_4_value = self.clean_value(extra_cell_4.value)
            extra_cell_5_value = self.clean_value(extra_cell_5.value)

            transformed_rng_cell_1 = transformed_values[0]
            transformed_rng_cell_2 = transformed_values[1]
            transformed_rng_cell_3 = transformed_values[2]

            # Use the updated match_bank_name logic with extra_cell_3_value and extra_cell_4_value
            bank_name = self.match_bank_name(extra_cell_3_value, extra_cell_4_value)

            # Calculate the equivalent in GBP based on "C:\\IT project3\\utils\\currencies.xlsx" currencies sheetname, which has ratios of all currencies against GBP.
            # it should be calculated for each row based on "extra_cell_1" (amount) and "extra_cell_2" (currency). The result should be saved in the "equivalent in GBP" column.
            # if the currencies don't match with the ones from the currencies.xlsx file, the "equivalent in GBP" column should be empty.
            # the structure of currencies.xlsx is as follows: columns; A =Currency Full Name, B = Currency Abbreviation, C= Exchange Ratio (1 unit = ? GBP),	D = 1 GBP equals (units)

            # Initialize the equivalent in GBP to an empty value.
            equivalent_in_GBP_value = None

            # Convert the extra_cell_1_value (amount) to a float.
            try:
                numeric_amount = float(extra_cell_1_value)
            except (ValueError, TypeError):
                print(f"Unable to convert amount '{extra_cell_1_value}' to a numeric value.")
                numeric_amount = None

            # Proceed only if the numeric amount is valid and a currency abbreviation is provided.
            if numeric_amount is not None and extra_cell_2_value is not None:
                # Define the path to the currencies file.
                # Ensure this path is correct. If you're running this on Windows, you can use raw strings.
                currency_file = os.path.join("C:\\IT project3\\utils", "currencies.xlsx")

                # Load the workbook and select the active worksheet.
                wb_currency = load_workbook(filename=currency_file)
                ws_currency = wb_currency.active

                # Iterate through the rows of the currencies file, starting from row 2 to skip the header.
                for row in ws_currency.iter_rows(min_row=2, values_only=True):
                    # Assuming the columns are as follows:
                    #   row[0] = Currency Full Name,
                    #   row[1] = Currency Abbreviation,
                    #   row[2] = Exchange Ratio (1 unit of the currency = ? GBP)
                    if row[1] == extra_cell_2_value:
                        # Convert the exchange ratio to a float, if it's not already a numeric type.
                        try:
                            exchange_ratio = float(row[2])
                        except (ValueError, TypeError):
                            print(
                                f"Invalid exchange ratio '{row[2]}' for currency '{extra_cell_2_value}' in currencies.xlsx.")
                            exchange_ratio = None

                        if exchange_ratio is not None:
                            equivalent_in_GBP_value = numeric_amount * exchange_ratio
                            # round to 2 decimal places
                            equivalent_in_GBP_value = round(equivalent_in_GBP_value, 2)
                        break  # Exit the loop once a match is found.

            # If no matching currency was found or if the conversion failed, handle accordingly.
            if extra_cell_2_value is not None and equivalent_in_GBP_value is None:
                # Assuming excel_file is defined elsewhere in your code, replace with the correct variable if needed.
                print(
                    f"Currency '{extra_cell_2_value}' not found or invalid exchange ratio in currencies.xlsx for case '{transformed_rng_cell_2}'.")
                equivalent_in_GBP_value = ""

            # At this point, equivalent_in_GBP_value holds the computed GBP value or remains empty.
            # print("Equivalent in GBP:", equivalent_in_GBP_value)

            values_excel_files[excel_file.name] = [
                transformed_rng_cell_1, transformed_rng_cell_2, extra_cell_1_value, extra_cell_2_value,
                extra_cell_3_value, extra_cell_4_value, extra_cell_5_value, bank_name, transformed_rng_cell_3,
                equivalent_in_GBP_value,
            ]
        return values_excel_files

    def create_combined_excel(self, values_excel_files):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet"

        header_list = ["uni name", "candidate name", "case nr", "amount", "currency", "acc number", "iban number",
                       "swift/bic", "name in bank", "vendor name", "equivalent in GBP"]

        # Write headers and data
        for i, header in enumerate(header_list):
            worksheet[f"A{i + 1}"] = header

        for i, excel_file in enumerate(values_excel_files):
            column_letter = get_column_letter(i + 2)
            worksheet[f"{column_letter}1"] = excel_file
            for j, value in enumerate(values_excel_files[excel_file]):
                if j == len(values_excel_files[excel_file]) - 3:
                    if value is not None and not isinstance(value, int):
                        value = str(value).replace("\n", "").replace("\xa0", "").replace("\t", "")
                worksheet[f"{column_letter}{j + 2}"] = value

        # Calculate sum
        row_index = 11
        sum_equivalent_in_GBP = sum(cell.value for cell in worksheet[row_index]
                                    if isinstance(cell.value, (int, float)))
        # convert to 2 decimal places
        sum_equivalent_in_GBP = round(sum_equivalent_in_GBP, 2)

        # Save workbook
        workbook.save(self.combined_file)

        # Create and process transposed sheet
        transposer = ExcelTransposer(self.combined_file)
        transposer.transpose_cells_to_table()

        # Add total to column L
        workbook = load_workbook(self.combined_file)
        worksheet_transposed = workbook["Transposed"]
        worksheet_transposed["L1"] = "Total in GBP:"
        worksheet_transposed["L2"] = sum_equivalent_in_GBP

        # create new sheet named "totals"
        worksheet_totals = workbook.create_sheet(title="Totals")
        # copy columns from "Transposed" sheet to "Totals" sheet - column B from Transposed to column A in Totals, D to B, E to C, K to D, L1 to E1, L2 to E2
        for i in range(1, worksheet_transposed.max_row + 1):
            worksheet_totals[f"A{i}"] = worksheet_transposed[f"B{i}"].value
            worksheet_totals[f"B{i}"] = worksheet_transposed[f"D{i}"].value
            worksheet_totals[f"C{i}"] = worksheet_transposed[f"E{i}"].value
            worksheet_totals[f"D{i}"] = worksheet_transposed[f"K{i}"].value
            worksheet_totals[f"E{i}"] = worksheet_transposed[f"L{i}"].value
        # save the workbook

        workbook.save(self.combined_file)

    def adjust_columns(self):
        # transposer = ExcelTransposer(self.combined_file)
        # transposer.transpose_cells_to_table()
        wide = Wide(self.combined_file, "Transposed", self.source_dir)
        wide.auto_adjust_column_width()

    def process_case_list(self):
        case_list = CaseList(self.source_dir, self.case_list_folder)
        case_list.process_excel_files()

    def compare_files(self):
        comparator = ExcelComparator(self.combined_file, self.comparison_file, self.source_dir)
        comparator.compare_and_append()

    def run(self):
        self.process_files()
        self.print_mismatched_cases()
        values_excel_files = self.collect_values()
        self.create_combined_excel(values_excel_files)
        self.adjust_columns()
        self.process_case_list()
        self.compare_files()


class TextRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.queue = queue.Queue()

    def write(self, message):
        self.queue.put(message)

    def flush(self):
        pass

    def clear(self):
        self.queue = queue.Queue()


class GUIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processing Hub")
        self.root.geometry("400x450")
        self.root.minsize(400, 450)
        self.root.maxsize(800, 600)

        frame = Frame(root)
        frame.pack(pady=20)

        self.label = Label(frame, text="Excel Processing Hub", font=("Arial", 18))
        self.label.pack(pady=10)
        self.label.bind("<Button-1>", self.check_hidden_popup)

        self.click_counter = 0
        self.click_time = 0

        self.mark_as_read_var = BooleanVar()
        self.mark_as_read_checkbox = Checkbutton(frame, text="Mark Emails as Read", variable=self.mark_as_read_var)
        self.mark_as_read_checkbox.pack(pady=5)

        self.email_button = Button(frame, text="Download Email Attachments", command=self.run_email_dl_task)
        self.email_button.pack(pady=5)

        self.kejsar_button = Button(frame, text="Process Cases", command=self.run_kejsar_task)
        self.kejsar_button.pack(pady=5)

        self.notes_button = Button(frame, text="Notes", command=self.create_notes_window)
        self.notes_button.pack(pady=5)

        self.instructions_button = Button(frame, text="Instructions", command=self.create_instructions_window)
        self.instructions_button.pack(pady=5)

        self.rename_button = Button(frame, text="Rename and Move Folder", command=self.rename_and_move_folder)
        self.rename_button.pack(pady=5)

        self.debug_text = Text(frame, wrap="word", height=10, width=50, state='disabled')
        self.debug_text.pack(pady=5)

        self.debug_output = TextRedirector(self.debug_text)
        sys.stdout = self.debug_output
        sys.stderr = self.debug_output

        self.update_debug_text()

        # Automatically open instructions window after 1 second when the app starts
        self.root.after(1000, self.create_instructions_window)

    def update_debug_text(self):
        while not self.debug_output.queue.empty():
            message = self.debug_output.queue.get()
            self.debug_text.config(state=tk.NORMAL)
            self.debug_text.insert("end", message)
            self.debug_text.see("end")
            self.debug_text.config(state=tk.DISABLED)
        self.root.after(500, self.update_debug_text)

    def update_gui_from_queue(self, q, text_widget):
        try:
            while not q.empty():
                message_type, message = q.get_nowait()
                if message_type == 'output' and message:
                    text_widget.config(state=tk.NORMAL)
                    text_widget.insert(tk.END, message)
                    text_widget.see(tk.END)  # Scroll to the end to show the latest message
                    text_widget.config(state=tk.DISABLED)
                elif message_type == 'error' and message:
                    text_widget.config(state=tk.NORMAL)
                    text_widget.insert(tk.END, "Error: " + message)
                    text_widget.see(tk.END)  # Scroll to the end to show the latest message
                    text_widget.config(state=tk.DISABLED)
                elif message_type == 'done' and message:
                    text_widget.config(state=tk.NORMAL)
                    text_widget.insert(tk.END, message)
                    text_widget.see(tk.END)  # Scroll to the end to show the latest message
                    text_widget.config(state=tk.DISABLED)
        except queue.Empty:
            pass
        finally:
            text_widget.after(1000, lambda: self.update_gui_from_queue(q, text_widget))  # Check the queue every second

    def print(self, message):
        self.debug_output += message + "\n"
        print(message)
        self.update_debug_text()

    def run_long_running_task(self, task_func, *args):
        task_thread = threading.Thread(target=task_func, args=args, daemon=True)
        task_thread.start()

    def run_email_dl_task(self):
        q = queue.Queue()
        self.run_long_running_task(self.run_email_dl, q)
        self.update_gui_from_queue(q, self.debug_text)

    def run_kejsar_task(self):
        q = queue.Queue()
        self.run_long_running_task(self.run_kejsar, q)
        self.update_gui_from_queue(q, self.debug_text)

    def import_function(self):
        pass
        # include the import function here

    def run_email_dl(self, q=None):
        threading.Thread(target=self._run_email_dl, args=(q,)).start()

    def _run_email_dl(self, q=None):
        # Initialize COM library in this thread
        pythoncom.CoInitialize()

        try:
            # Get the absolute path of the script
            script_path = os.path.dirname(os.path.abspath(__file__))

            # Path to save attachments
            attachment_save_path = os.path.join(script_path, 'outlook/excel')

            # Path to save Outlook messages (.msg)
            msg_save_path = os.path.join(script_path, 'outlook/msg')

            # Create directories if they don't exist
            os.makedirs(attachment_save_path, exist_ok=True)
            os.makedirs(msg_save_path, exist_ok=True)

            if attachment_save_path and msg_save_path:
                category_to_download = ''  # Change as needed
                target_senders = ['emails to be input here']
                # Fill with target sender emails
                # /O=EXCHANGELABS/OU=EXCHANGE ADMINISTRATIVE GROUP (FYDIBOHF23SPDLT)/CN=RECIPIENTS/CN=A85096D0D7D248ED9BCEB267934F4D96-8042407E-08 <<< Anil
                processor = OutlookProcessor(category_to_download, target_senders, attachment_save_path, msg_save_path)
                initial_file_count = processor.count_files_in_directory(attachment_save_path)
                num_unread_emails = processor.list_unread_emails()
                print(f"Initial file count: {initial_file_count}, Unread emails: {num_unread_emails}")
                # ask the user if all previous files found within outlook folder should be deleted if the amount is more than 0. If yes, delete all files from the folder, if not, keep them
                if initial_file_count > 0:
                    messagebox.showinfo("Info", "There are files in the outlook folder. Do you want to delete them?")
                    if messagebox.askyesno("Delete Files", "Do you want to delete the files?"):
                        for file in os.listdir(attachment_save_path):
                            file_path = os.path.join(attachment_save_path, file)
                            try:
                                if os.path.isfile(file_path):
                                    os.unlink(file_path)
                            except Exception as e:
                                print(f"Error deleting file: {str(e)}")
                        print("All files deleted from the outlook folder.")
                    else:
                        print("Files were not deleted from the outlook folder.")

                if num_unread_emails > 0:
                    processor.download_attachments_and_save_as_msg(True, self.mark_as_read_var.get())
                else:
                    print("No unread emails to process.")
                messagebox.showinfo("Info", "Email processing completed.")
        finally:
            # Uninitialize COM library in this thread
            pythoncom.CoUninitialize()

    def run_kejsar(self, q):
        # Get the absolute path of the script
        script_path = os.path.dirname(os.path.abspath(__file__))
        processor = KejsarProcessor(script_path)
        processor.run()
        messagebox.showinfo("Info", "Kejsar processing completed.")

    def create_notes_window(self):
        global notes_window, notes_content, notes_first_time
        if not notes_window or not tk.Toplevel.winfo_exists(notes_window):
            notes_window = tk.Toplevel()
            notes_window.title("Notes")
            notes_window.geometry("300x200+200+50")
            notes_window.resizable(True, True)
            notes_window.minsize(150, 100)
            notes_window.maxsize(600, 400)
            text_widget = tk.Text(notes_window, wrap=tk.WORD)
            text_widget.grid(row=0, column=0, sticky='nsew')
            notes_window.grid_rowconfigure(0, weight=1)
            notes_window.grid_columnconfigure(0, weight=1)
            text_widget.insert(tk.END, notes_content)

            def on_closing():
                global notes_window, notes_content
                notes_content = text_widget.get("1.0", tk.END)
                notes_window.destroy()
                notes_window = None

            notes_window.bind("<Escape>", lambda event: on_closing())
            notes_window.protocol("WM_DELETE_WINDOW", on_closing)
            notes_window.focus_set()

            if notes_first_time:
                notes_window.after(1000, lambda: messagebox.showinfo("Note",
                                                                     "Please note that any notes you make will disappear when closing the program."))
                notes_first_time = False

    def create_instructions_window(self):
        global instructions_window
        if not instructions_window or not tk.Toplevel.winfo_exists(instructions_window):
            instructions_window = tk.Toplevel()
            instructions_window.title("Instructions")
            instructions_window.geometry("800x500")
            text_widget = tk.Text(instructions_window, height=500, width=500, wrap=tk.WORD)
            text_widget.pack(padx=10, pady=10)
            text_widget.insert(tk.END,
                               "Instructions:\n\n - Click the 'Download Email Attachments' button to process emails. Decide if you want to mark them as read or not.\n\n - Click the 'Process Cases' button to process cases. This will exctract all the saved excel files and create a new 'combined_file' excel file on which you can work.\n\n - Click the 'Notes' button to view or edit your personal notes. Or use notepad idk.\n\n - Click the 'Rename and Move Folder' button to rename and move the folder with the cases to an archive folder. \n\n This standalone app might be saving files in weird places, so be sure to check the printouts for the paths.")
            text_widget.config(state=tk.DISABLED)

            # open instructions window automatically 1 second after the GUI opens on start

            # Bind the Escape key to close the window
            def on_closing():
                global instructions_window
                instructions_window.destroy()
                instructions_window = None

            instructions_window.bind("<Escape>", lambda event: on_closing())

            instructions_window.protocol("WM_DELETE_WINDOW", on_closing)

            instructions_window.focus_set()

    def check_hidden_popup(self, event):
        current_time = time.time()
        if current_time - self.click_time > 3:
            self.click_counter = 0
        self.click_time = current_time
        self.click_counter += 1
        if self.click_counter == 7:
            self.show_hidden_popup()
            self.click_counter = 0

    def show_hidden_popup(self):
        hidden_popup = tk.Toplevel(self.root)
        hidden_popup.title("Hidden Section")
        hidden_popup.geometry("100x60")
        hidden_popup.resizable(False, False)

        import_button = Button(hidden_popup, text="Import", command=self.import_function)
        import_button.pack(padx=10, pady=10)

    def rename_and_move_folder(self):
        try:
            # Define paths
            source_dir = r"C:\\IT project3\\utils\\pmt_run"
            dest_dir = r"C:\\IT project3\\prev runs\\2k25"
            new_folder_name = datetime.now().strftime("%d.%m")
            new_folder_path = os.path.join(dest_dir, new_folder_name)

            # Check if the source folder exists
            if not os.path.exists(source_dir):
                print(f"Source folder not found: {source_dir}")
                return

            # Check if destination folder exists, if not create it
            if not os.path.exists(dest_dir):
                os.makedirs(dest_dir)

            # Rename and move the folder
            shutil.move(source_dir, new_folder_path)
            print(f"Folder renamed to {new_folder_name} and moved to {dest_dir}")

            # Create a new "pmt_run" folder in the utils directory
            os.makedirs(source_dir)
            print(f"New 'pmt_run' folder created in {source_dir}")

        except Exception as e:
            print(f"Error: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = GUIApp(root)
    root.mainloop()

# update GUI so that each email is shown in a prinout section on its own rather than visible details of two emails at once
# match the filename with the university provided inside the excel file