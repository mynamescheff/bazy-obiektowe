import os
import re
import time
import shutil
import datetime
from datetime import date, datetime
from pathlib import Path
import tkinter as tk
from tkinter import Tk, Frame, Label, Button, Entry, Text, Toplevel, filedialog, messagebox, Scrollbar, RIGHT, Y, Checkbutton, BooleanVar, W, E
from tkinter import ttk
import win32com.client
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import threading
import sys

# Replace with actual email or create input field for it
SHARED_MAILBOX_EMAIL = ""  

class CaseList:
    def __init__(self, excel_folder, list_folder):
        self.excel_folder = excel_folder
        self.list_folder = list_folder

    def process_excel_files(self):
        list_file_path = os.path.join(self.list_folder, "case_list.txt")
        existing_values = {}
        duplicate_counts = {}
        error_messages = []
        
        if os.path.isfile(list_file_path):
            existing_values, duplicate_counts = self.load_existing_list(list_file_path)
            
        all_entries = []
        for file_name in os.listdir(self.excel_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(self.excel_folder, file_name)
                try:
                    wb = load_workbook(file_path)
                    sheet = wb.active
                    value = sheet["G2"].value
                    if value:
                        # Clean the value
                        value = self._clean_string(value)
                        
                        if value in existing_values:
                            duplicate_counts[value] = duplicate_counts.get(value, 0) + 1
                            entry = f"{value} [{file_name} - DUPLICATE {duplicate_counts[value]}]"
                            all_entries.append(entry)
                            print(f"Duplicate found: {value} in file {file_name} (Duplicate count: {duplicate_counts[value]})")
                        else:
                            existing_values[value] = False
                            duplicate_counts[value] = 0
                            entry = f"{value} [{file_name}]"
                            all_entries.append(entry)
                except Exception as e:
                    error_messages.append(f"Error processing file '{file_name}': {str(e)}")
                    
        if all_entries:
            today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Ensure the directory exists
            os.makedirs(os.path.dirname(list_file_path), exist_ok=True)
            
            # Create the file if it doesn't exist
            if not os.path.exists(list_file_path):
                with open(list_file_path, "w", encoding="utf-8") as file:
                    file.write(f"--- Case List Created on {today} ---\n")
            
            # Append the new entries
            with open(list_file_path, "a", encoding="utf-8") as file:
                file.write(f"\n--- Updated on {today} ---\n")
                for entry in all_entries:
                    file.write(f"{entry} ({today})\n")
                            
                print("Cases processed successfully and saved to the case_list.txt file.")
                return duplicate_counts, error_messages

    def _clean_string(self, value):
        if not value:
            return ""
            
        value = str(value)
        value = re.sub(r'[\n\r\t\v\f\x85\u2028\u2029]+', ' ', value)
        
        # Remove extra spaces
        while "  " in value:
            value = value.replace("  ", " ")
            
        # Make sure string ends with alphanumeric character
        while value and not value[-1].isalnum():
            value = value[:-1]
            
        return value.strip()

    def load_existing_list(self, list_file_path):
        existing_values = {}
        duplicate_counts = {}
        with open(list_file_path, "r", encoding="utf-8") as file:
            for line in file:
                if line.strip() and not line.startswith("---"):
                    parts = line.split(" [")
                    if len(parts) > 1:
                        value = parts[0]
                        existing_values[value] = False
                        duplicate_count = parts[1].count("DUPLICATE")
                        duplicate_counts[value] = duplicate_count
        return existing_values, duplicate_counts

class OutlookProcessor:
    def __init__(self, category, target_senders, attachment_save_path, msg_save_path):
        self.category = category
        self.target_senders = target_senders
        self.attachment_save_path = attachment_save_path
        self.msg_save_path = msg_save_path
        self.processed_emails = {}
        self.emails_with_pdf = []
        self.emails_with_nvf_new_vendor = []
        
        # Create directories if they don't exist
        os.makedirs(self.attachment_save_path, exist_ok=True)
        os.makedirs(self.msg_save_path, exist_ok=True)

    def initialize_outlook(self):
        try:
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            self.namespace = self.outlook.GetNamespace("MAPI")
            return True
        except Exception as e:
            print(f"Error initializing Outlook: {str(e)}")
            return False

    def list_unread_emails(self):
        if not hasattr(self, 'outlook') or not hasattr(self, 'namespace'):
            if not self.initialize_outlook():
                return 0
                
        recipient = self.namespace.CreateRecipient(SHARED_MAILBOX_EMAIL)
        recipient.Resolve()
        if recipient.Resolved:
            shared_mailbox = self.namespace.GetSharedDefaultFolder(recipient, 6)
            unread_emails = shared_mailbox.Items.Restrict(f"[Categories] = '{self.category}' AND [UnRead] = True")
            return len([email for email in unread_emails])
        return 0

    def get_unique_filename(self, base_path, original_filename, extension):
        counter = 2
        new_filename = original_filename
        while os.path.exists(os.path.join(base_path, f"{new_filename}{extension}")):
            new_filename = f"{original_filename} {counter}"
            counter += 1
        return new_filename

    def mark_email_as_read(self, item, mark_as_read):
        item.UnRead = not mark_as_read
        item.Save()
        if mark_as_read:
            print("Marked email as read.")
        else:
            print("Email left as unread.")

    def transform_to_swift_accepted_characters(self, text_list):
        """Basic character transformation - simplified from what was referenced"""
        if not text_list:
            return []
            
        result = []
        for text in text_list:
            if text:
                # Remove potentially problematic characters
                clean_text = re.sub(r'[\/:*?"<>|\t]', ' ', str(text))
                result.append(clean_text)
            else:
                result.append("")
        return result

    def download_attachments_and_save_as_msg(self, save_emails, mark_as_read):
        if not hasattr(self, 'outlook') or not hasattr(self, 'namespace'):
            if not self.initialize_outlook():
                print("Failed to initialize Outlook. Cannot process emails.")
                return
                
        recipient = self.namespace.CreateRecipient(SHARED_MAILBOX_EMAIL)
        recipient.Resolve()
        self.processed_emails = {}  # Clear previous tracking
        self.emails_with_pdf = []  # Clear previous tracking
        self.emails_with_nvf_new_vendor = []  # Clear previous tracking
        emails_with_no_attachments = []  # Track emails with no attachments
        
        if recipient.Resolved:
            shared_mailbox = self.namespace.GetSharedDefaultFolder(recipient, 6)
            unread_emails = shared_mailbox.Items.Restrict(f"[Categories] = '{self.category}' AND [UnRead] = True")
            emails_to_process = [email for email in unread_emails]
            print(f"Found {len(emails_to_process)} emails to process under category '{self.category}' with 'UnRead' = True.")
            
            if save_emails:
                saved_emails = 0
                saved_attachments = 0
                not_saved_subjects = []
                incorrect_subjects = []
                
                for item in emails_to_process:
                    time.sleep(2)
                    try:
                        sender_email = item.SenderEmailAddress
                        sender_name_match = re.search(r'/O=EXCHANGELABS/OU=EXCHANGE ADMINISTRATIVE GROUP.*?-([A-Za-z]+)', sender_email)
                        sender_name = sender_name_match.group(1) if sender_name_match else sender_email
                        print(f"Processing email from: {sender_name}")
                        
                        if sender_name.lower() in [sender.lower() for sender in self.target_senders]:
                            if item.Attachments.Count == 0:
                                # Email has no attachments
                                emails_with_no_attachments.append(item.Subject)
                                print(f"Email with subject '{item.Subject}' has no attachments and will not be processed.")
                                continue  # Skip to the next email

                            # Process email with attachments
                            subject_correct = True
                            saved_attachment_paths = []
                            has_pdf = False
                            has_nvf_new_vendor = False
                            save_msg_once = False
                            approval_msg_path = ""
                            
                            if item.Attachments.Count > 0:
                                for attachment in item.Attachments:
                                    if attachment.FileName.lower().endswith('.pdf'):
                                        has_pdf = True
                                    if 'nvf' in attachment.FileName.lower() or 'new vendor' in attachment.FileName.lower() or 'vendor' in attachment.FileName.lower():
                                        has_nvf_new_vendor = True
                                    if attachment.FileName.lower().endswith('.xlsx'):
                                        new_filename = self.extract_filename_from_subject(item.Subject)
                                        new_filename = self.transform_to_swift_accepted_characters([new_filename])[0]
                                        
                                        if ';' not in new_filename and ';' not in item.Subject:
                                            print(f"Invalid filename generated from subject: {new_filename}")
                                            subject_correct = False
                                            incorrect_subjects.append(item.Subject)
                                            continue
                                            
                                        unique_attachment_filename = self.get_unique_filename(self.attachment_save_path, new_filename, '.xlsx')
                                        attachment_path = os.path.join(self.attachment_save_path, f"{unique_attachment_filename}.xlsx")
                                        attachment.SaveAsFile(attachment_path)
                                        print(f"Saved attachment: {attachment_path}")
                                        saved_attachment_paths.append(attachment_path)
                                        saved_attachments += 1
                                        
                                        if not save_msg_once:
                                            unique_msg_filename = self.get_unique_filename(self.msg_save_path, new_filename, ' approval.msg')
                                            approval_msg_path = os.path.join(self.msg_save_path, f"{unique_msg_filename} approval.msg")
                                            item.SaveAs(approval_msg_path)
                                            print(f"Saved Outlook message: {approval_msg_path}")
                                            save_msg_once = True
                                            
                                if has_pdf:
                                    self.emails_with_pdf.append(item.Subject)
                                if has_nvf_new_vendor:
                                    self.emails_with_nvf_new_vendor.append(item.Subject)
                                    print(f"NVF is present in the email with subject: {item.Subject}")
                                    
                                if subject_correct and not has_nvf_new_vendor:
                                    self.mark_email_as_read(item, mark_as_read)
                                    saved_emails += 1
                                    self.processed_emails[item.Subject] = saved_attachment_paths
                                elif has_nvf_new_vendor and approval_msg_path:
                                    self.clean_up_files(saved_attachment_paths, approval_msg_path)
                                    
                    except Exception as e:
                        print(f"Error processing email: {str(e)}")
                        not_saved_subjects.append(item.Subject)
                        
                print(f"Total saved emails: {saved_emails}")
                actual_saved_attachments = self.count_files_in_directory(self.attachment_save_path)
                print(f"Total saved attachments: {actual_saved_attachments}")
                
                if not_saved_subjects:
                    print("Emails not saved due to errors:")
                    for subject in not_saved_subjects:
                        print(subject)
                        
                if incorrect_subjects:
                    print("Emails with incorrect subject format:")
                    for subject in incorrect_subjects:
                        print(subject)
                        
                # Copy attachments to the utils/pmt_run directory
                self.copy_attachments_to_pmt_run()

                # Create log file
                self.create_log_file(emails_with_no_attachments, not_saved_subjects)

            else:
                print("Email saving is disabled. No emails were saved.")
        else:
            print(f"Could not resolve the recipient: {SHARED_MAILBOX_EMAIL}")

    def create_log_file(self, emails_with_no_attachments, not_saved_subjects):
        log_directory = r"C:\IT project3\utils\logs"
        os.makedirs(log_directory, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"process_log_{timestamp}.txt"
        log_path = os.path.join(log_directory, log_filename)

        with open(log_path, 'w', encoding='utf-8') as log_file:
            log_file.write("Summary of processed emails and saved attachments:\n")
            if self.processed_emails:
                for subject, attachments in self.processed_emails.items():
                    # Transform the subject
                    transformed_subject = self.transform_to_swift_accepted_characters([subject])[0]
                    log_file.write(f"Email subject: {transformed_subject}\n")
                    # Transform the attachment names
                    transformed_attachments = self.transform_to_swift_accepted_characters(attachments)
                    for attachment in transformed_attachments:
                        log_file.write(f" - Saved attachment: {attachment}\n")

            actual_saved_attachments = self.count_files_in_directory(self.attachment_save_path)
            expected_attachments = sum(len(paths) for paths in self.processed_emails.values())
            if expected_attachments != actual_saved_attachments:
                log_file.write(f"Discrepancy detected: expected {expected_attachments} attachments, but found {actual_saved_attachments} in directory.\n")

            if self.emails_with_pdf:
                log_file.write("Emails with PDF attachments:\n")
                for subject in self.emails_with_pdf:
                    log_file.write(f"{subject}\n")

            if self.emails_with_nvf_new_vendor:
                log_file.write("Emails with NVF or New Vendor in attachment filenames:\n")
                for subject in self.emails_with_nvf_new_vendor:
                    log_file.write(f"{subject}\n")

            if emails_with_no_attachments:
                log_file.write("Emails with no attachments:\n")
                for subject in emails_with_no_attachments:
                    log_file.write(f"{subject}\n")

            if not_saved_subjects:
                log_file.write("Emails not saved due to errors:\n")
                for subject in not_saved_subjects:
                    log_file.write(f"{subject}\n")
                    
        print(f"Log file created at {log_path}")

    def clean_up_files(self, attachment_paths, msg_path):
        for attachment_path in attachment_paths:
            try:
                os.remove(attachment_path)
                print(f"Deleted attachment: {attachment_path}")
            except Exception as e:
                print(f"Error deleting attachment {attachment_path}: {str(e)}")
        try:
            os.remove(msg_path)
            print(f"Deleted Outlook message: {msg_path}")
        except Exception as e:
            print(f"Error deleting Outlook message {msg_path}: {str(e)}")

    def copy_attachments_to_pmt_run(self):
        # Get the absolute path of the script
        script_path = os.path.dirname(os.path.abspath(__file__))
        # Define the pmt_run directory relative to the script location
        pmt_run_dir = os.path.join(script_path, "utils", "pmt_run")
        os.makedirs(pmt_run_dir, exist_ok=True)
        
        for filename in os.listdir(self.attachment_save_path):
            source_file = os.path.join(self.attachment_save_path, filename)
            destination_file = os.path.join(pmt_run_dir, filename)
            shutil.copy(source_file, destination_file)
            
        print(f"All files copied to {pmt_run_dir}")

    def extract_filename_from_subject(self, subject):
        match = re.search(r';\s*(.*)', subject)
        if match:
            return match.group(1)
        else:
            return subject

    def count_files_in_directory(self, directory):
        return len([name for name in os.listdir(directory) if os.path.isfile(os.path.join(directory, name))])

class ExcelTransposer:
    def __init__(self, filename: str):
        self.filename = filename
        self.workbook: Workbook = load_workbook(filename)
        self.sheet: Worksheet = self.workbook.active

    def set_active_sheet(self, sheet_name: str) -> None:
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    def transpose_cells_to_table(self) -> None:
        try:
            data = []
            for row in self.sheet.iter_rows(values_only=True):
                data.append(row)
            transposed_data = list(map(list, zip(*data)))
            transposed_sheet = self.workbook.create_sheet(title="Transposed")
            
            for row_idx, row_data in enumerate(transposed_data):
                for col_idx, cell_value in enumerate(row_data):
                    column_letter = get_column_letter(col_idx + 1)
                    transposed_sheet[f"{column_letter}{row_idx + 1}"] = cell_value
                    
            self.auto_adjust_column_width(transposed_sheet)
            self.workbook.save(self.filename)
        except Exception as e:
            print(f"An error occurred while transposing the data: {e}")

    def auto_adjust_column_width(self, sheet: Worksheet) -> None:
        for column in sheet.columns:
            max_length = 0
            column_cells = [cell for cell in column]
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

# New class for scraping Excel values
class ExcelDataScraper:
    def __init__(self, directory=None):
        self.directory = directory
        self.results = []
        self.headers = []

    def set_directory(self, directory):
        self.directory = directory
        
    def scrape_excel_files(self, range_start="G2", range_end="G2", read_headers=True):
        """
        Scrape values from Excel files in the specified directory.
        Default is now G2 cell specifically.
        """
        if not self.directory or not os.path.isdir(self.directory):
            raise ValueError("Please set a valid directory first.")
            
        self.results = []
        self.headers = []
        
        # Extract column letters from range
        start_col = range_start[0]
        end_col = range_end[0]
        
        # Extract row numbers from range
        start_row = int(range_start[1:])
        end_row = int(range_end[1:])
        
        # Convert column letters to indices
        start_col_idx = ord(start_col) - ord('A') + 1
        end_col_idx = ord(end_col) - ord('A') + 1
        
        for filename in os.listdir(self.directory):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(self.directory, filename)
                try:
                    workbook = load_workbook(file_path, data_only=True)
                    sheet = workbook.active
                    
                    # Read headers if specified
                    if read_headers and not self.headers:
                        header_row = []
                        for col_idx in range(start_col_idx, end_col_idx + 1):
                            col_letter = get_column_letter(col_idx)
                            cell_value = sheet[f"{col_letter}1"].value
                            header_row.append(cell_value if cell_value else f"Column {col_letter}")
                        self.headers = header_row
                    
                    # Read data from specified range
                    file_data = {"filename": filename, "values": {}}
                    for row in range(start_row, end_row + 1):
                        row_data = []
                        for col_idx in range(start_col_idx, end_col_idx + 1):
                            col_letter = get_column_letter(col_idx)
                            cell_value = sheet[f"{col_letter}{row}"].value
                            row_data.append(cell_value)
                            
                        # Map data to headers if available
                        if self.headers:
                            for idx, header in enumerate(self.headers):
                                if idx < len(row_data):
                                    file_data["values"][header] = row_data[idx]
                        else:
                            # Use column letters as keys if no headers
                            for idx, value in enumerate(row_data):
                                col_letter = get_column_letter(start_col_idx + idx)
                                file_data["values"][col_letter] = value
                    
                    self.results.append(file_data)
                except Exception as e:
                    print(f"Error processing {filename}: {str(e)}")
                    
        return self.results
    
    def get_headers(self):
        return self.headers
    
    def get_results(self):
        return self.results
    
    def save_results_to_csv(self, output_file):
        """Save scraped results to a CSV file"""
        if not self.results:
            print("No results to save.")
            return False
            
        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                # Determine headers
                if self.headers:
                    fieldnames = ['filename'] + self.headers
                else:
                    # Use the keys from the first result's values
                    fieldnames = ['filename'] + list(self.results[0]['values'].keys())
                
                import csv
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in self.results:
                    row = {'filename': result['filename']}
                    row.update(result['values'])
                    writer.writerow(row)
                    
            print(f"Results saved to {output_file}")
            return True
        except Exception as e:
            print(f"Error saving results: {str(e)}")
            return False
        
    def save_results_to_excel(self, output_file):
        """Save scraped results to an Excel file"""
        if not self.results:
            print("No results to save.")
            return False
            
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Combined Data"
            
            # Write headers
            if self.headers:
                headers = ['filename'] + self.headers
            else:
                # Use the keys from the first result's values
                headers = ['filename'] + list(self.results[0]['values'].keys())
                
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Write data
            for row_idx, result in enumerate(self.results, 2):
                ws.cell(row=row_idx, column=1, value=result['filename'])
                
                for col_idx, header in enumerate(headers[1:], 2):
                    value = result['values'].get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
                    
            wb.save(output_file)
            print(f"Results saved to Excel file: {output_file}")
            return True
        except Exception as e:
            print(f"Error saving results to Excel: {str(e)}")
            return False

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Processor Tool")
        self.root.geometry("800x600")
        
        # Create a notebook with tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create frames for each tab
        self.outlook_tab = ttk.Frame(self.notebook)
        self.case_list_tab = ttk.Frame(self.notebook)
        self.transpose_tab = ttk.Frame(self.notebook)
        self.scrape_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.outlook_tab, text="Outlook Processor")
        self.notebook.add(self.case_list_tab, text="Case List")
        self.notebook.add(self.transpose_tab, text="Excel Transpose")
        self.notebook.add(self.scrape_tab, text="Excel Scraper")
        
        # Setup tabs
        self.setup_outlook_tab()
        self.setup_case_list_tab()
        self.setup_transpose_tab()
        self.setup_scrape_tab()
        
        # Initialize components
        self.excel_scraper = ExcelDataScraper()
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        
        # Status bar at the bottom
        status_frame = ttk.Frame(root)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=5)
        ttk.Label(status_frame, textvariable=self.status_var).pack(side=tk.LEFT)

    def export_to_excel(self):
        try:
            if not self.excel_scraper.get_results():
                messagebox.showerror("Error", "No data to export. Please scrape Excel files first.")
                return
                
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )
            
            if not output_file:
                return  # User cancelled
                
            success = self.excel_scraper.save_results_to_excel(output_file)
            
            if success:
                messagebox.showinfo("Success", f"Data exported to {output_file}")
                self.status_var.set("Data exported to Excel")
            else:
                messagebox.showerror("Error", "Failed to export data")
                
        except Exception as e:
            messagebox.showerror("Error", f"Export error: {str(e)}")
            self.status_var.set("Error exporting data")

    def setup_outlook_tab(self):
        # Create form for Outlook processor
        ttk.Label(self.outlook_tab, text="Email Category:").grid(row=0, column=0, sticky=W, padx=10, pady=5)
        self.category_entry = ttk.Entry(self.outlook_tab, width=40)
        self.category_entry.grid(row=0, column=1, sticky=W, padx=10, pady=5)
        self.category_entry.insert(0, "Approval")
        
        ttk.Label(self.outlook_tab, text="Target Senders (comma-separated):").grid(row=1, column=0, sticky=W, padx=10, pady=5)
        self.senders_entry = ttk.Entry(self.outlook_tab, width=40)
        self.senders_entry.grid(row=1, column=1, sticky=W, padx=10, pady=5)
        self.senders_entry.insert(0, "Sender1,Sender2")
        
        ttk.Label(self.outlook_tab, text="Attachments Save Path:").grid(row=2, column=0, sticky=W, padx=10, pady=5)
        self.attachment_path_frame = ttk.Frame(self.outlook_tab)
        self.attachment_path_frame.grid(row=2, column=1, sticky=W, padx=10, pady=5)
        self.attachment_path_entry = ttk.Entry(self.attachment_path_frame, width=30)
        self.attachment_path_entry.pack(side=tk.LEFT)
        self.attachment_path_entry.insert(0, "C:/Attachments")
        ttk.Button(self.attachment_path_frame, text="Browse", command=lambda: self.browse_directory(self.attachment_path_entry)).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(self.outlook_tab, text="Messages Save Path:").grid(row=3, column=0, sticky=W, padx=10, pady=5)
        self.msg_path_frame = ttk.Frame(self.outlook_tab)
        self.msg_path_frame.grid(row=3, column=1, sticky=W, padx=10, pady=5)
        self.msg_path_entry = ttk.Entry(self.msg_path_frame, width=30)
        self.msg_path_entry.pack(side=tk.LEFT)
        self.msg_path_entry.insert(0, "C:/Messages")
        ttk.Button(self.msg_path_frame, text="Browse", command=lambda: self.browse_directory(self.msg_path_entry)).pack(side=tk.LEFT, padx=5)
        
        self.mark_as_read_var = BooleanVar(value=True)
        ttk.Checkbutton(self.outlook_tab, text="Mark emails as read", variable=self.mark_as_read_var).grid(row=4, column=0, sticky=W, padx=10, pady=5)
        
        self.save_emails_var = BooleanVar(value=True)
        ttk.Checkbutton(self.outlook_tab, text="Save emails", variable=self.save_emails_var).grid(row=4, column=1, sticky=W, padx=10, pady=5)
        
        ttk.Button(self.outlook_tab, text="Check Unread Emails", command=self.check_unread_emails).grid(row=5, column=0, sticky=W, padx=10, pady=5)
        ttk.Button(self.outlook_tab, text="Process Emails", command=self.process_emails).grid(row=5, column=1, sticky=W, padx=10, pady=5)
        
        # Results display
        ttk.Label(self.outlook_tab, text="Processing Results:").grid(row=6, column=0, sticky=W, padx=10, pady=5)
        self.outlook_result_text = Text(self.outlook_tab, height=15, width=80)
        self.outlook_result_text.grid(row=7, column=0, columnspan=2, padx=10, pady=5)
        
        # Add scrollbar
        outlook_scrollbar = ttk.Scrollbar(self.outlook_tab, command=self.outlook_result_text.yview)
        outlook_scrollbar.grid(row=7, column=2, sticky='nsew')
        self.outlook_result_text.config(yscrollcommand=outlook_scrollbar.set)


    def setup_case_list_tab(self):
        ttk.Label(self.case_list_tab, text="Excel Files Folder:").grid(row=0, column=0, sticky=W, padx=10, pady=5)
        self.excel_folder_frame = ttk.Frame(self.case_list_tab)
        self.excel_folder_frame.grid(row=0, column=1, sticky=W, padx=10, pady=5)
        self.excel_folder_entry = ttk.Entry(self.excel_folder_frame, width=30)
        self.excel_folder_entry.pack(side=tk.LEFT)
        ttk.Button(self.excel_folder_frame, text="Browse", command=lambda: self.browse_directory(self.excel_folder_entry)).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(self.case_list_tab, text="List Output Folder:").grid(row=1, column=0, sticky=W, padx=10, pady=5)
        self.list_folder_frame = ttk.Frame(self.case_list_tab)
        self.list_folder_frame.grid(row=1, column=1, sticky=W, padx=10, pady=5)
        self.list_folder_entry = ttk.Entry(self.list_folder_frame, width=30)
        self.list_folder_entry.pack(side=tk.LEFT)
        ttk.Button(self.list_folder_frame, text="Browse", command=lambda: self.browse_directory(self.list_folder_entry)).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(self.case_list_tab, text="Process Case List", command=self.process_case_list).grid(row=2, column=0, columnspan=2, padx=10, pady=10)
        
        # Results display
        ttk.Label(self.case_list_tab, text="Processing Results:").grid(row=3, column=0, sticky=W, padx=10, pady=5)
        self.case_list_result_text = Text(self.case_list_tab, height=15, width=80)
        self.case_list_result_text.grid(row=4, column=0, columnspan=2, padx=10, pady=5)
        
        # Add scrollbar
        case_list_scrollbar = ttk.Scrollbar(self.case_list_tab, command=self.case_list_result_text.yview)
        case_list_scrollbar.grid(row=4, column=2, sticky='nsew')
        self.case_list_result_text.config(yscrollcommand=case_list_scrollbar.set)

    def setup_transpose_tab(self):
        ttk.Label(self.transpose_tab, text="Excel File:").grid(row=0, column=0, sticky=W, padx=10, pady=5)
        self.excel_file_frame = ttk.Frame(self.transpose_tab)
        self.excel_file_frame.grid(row=0, column=1, sticky=W, padx=10, pady=5)
        self.excel_file_entry = ttk.Entry(self.excel_file_frame, width=30)
        self.excel_file_entry.pack(side=tk.LEFT)
        ttk.Button(self.excel_file_frame, text="Browse", command=lambda: self.browse_file(self.excel_file_entry, [("Excel Files", "*.xlsx")])).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(self.transpose_tab, text="Sheet Name (optional):").grid(row=1, column=0, sticky=W, padx=10, pady=5)
        self.sheet_name_entry = ttk.Entry(self.transpose_tab, width=30)
        self.sheet_name_entry.grid(row=1, column=1, sticky=W, padx=10, pady=5)
        
        ttk.Button(self.transpose_tab, text="Transpose Excel", command=self.transpose_excel).grid(row=2, column=0, columnspan=2, padx=10, pady=10)
        
        # Results display
        ttk.Label(self.transpose_tab, text="Transpose Results:").grid(row=3, column=0, sticky=W, padx=10, pady=5)
        self.transpose_result_text = Text(self.transpose_tab, height=15, width=80)
        self.transpose_result_text.grid(row=4, column=0, columnspan=2, padx=10, pady=5)
    
    def setup_scrape_tab(self):
        ttk.Label(self.scrape_tab, text="Excel Files Directory:").grid(row=0, column=0, sticky=W, padx=10, pady=5)
        self.scrape_dir_frame = ttk.Frame(self.scrape_tab)
        self.scrape_dir_frame.grid(row=0, column=1, sticky=W, padx=10, pady=5)
        self.scrape_dir_entry = ttk.Entry(self.scrape_dir_frame, width=30)
        self.scrape_dir_entry.pack(side=tk.LEFT)
        ttk.Button(self.scrape_dir_frame, text="Browse", command=lambda: self.browse_directory(self.scrape_dir_entry)).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.scrape_dir_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        
        # Range selection
        range_frame = ttk.Frame(self.scrape_tab)
        range_frame.grid(row=1, column=0, columnspan=2, sticky=W, padx=10, pady=5)
        
        ttk.Label(range_frame, text="Cell Range:").pack(side=tk.LEFT, padx=5)
        ttk.Label(range_frame, text="From:").pack(side=tk.LEFT, padx=5)
        self.range_start_entry = ttk.Entry(range_frame, width=8)
        self.range_start_entry.pack(side=tk.LEFT, padx=5)
        self.range_start_entry.insert(0, "A2")
        ttk.Label(range_frame, text="To:").pack(side=tk.LEFT, padx=5)
        self.range_end_entry = ttk.Entry(range_frame, width=8)
        self.range_end_entry.pack(side=tk.LEFT, padx=5)
        self.range_end_entry.insert(0, "F2")
        
        # Read headers option
        self.read_headers_var = BooleanVar(value=True)
        ttk.Checkbutton(self.scrape_tab, text="Read headers from first row", variable=self.read_headers_var).grid(row=2, column=0, columnspan=2, sticky=W, padx=10, pady=5)
        
        # Buttons
        buttons_frame = ttk.Frame(self.scrape_tab)
        buttons_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=5)
        
        ttk.Button(buttons_frame, text="Scrape Excel Files", command=self.scrape_excel_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Export to CSV", command=self.export_to_csv).pack(side=tk.LEFT, padx=5)
        
        # Results display
        ttk.Label(self.scrape_tab, text="Scraped Data:").grid(row=4, column=0, sticky=W, padx=10, pady=5)
        
        
        # Create a Frame to contain the Text widget and scrollbars
        results_frame = ttk.Frame(self.scrape_tab)
        results_frame.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky='nsew')
        
        # Make the frame expandable
        self.scrape_tab.rowconfigure(5, weight=1)
        self.scrape_tab.columnconfigure(1, weight=1)
        
        # Create Text widget with scrollbars
        self.scrape_result_text = Text(results_frame, height=15, width=80)
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.scrape_result_text.yview)
        h_scrollbar = ttk.Scrollbar(results_frame, orient="horizontal", command=self.scrape_result_text.xview)
        
        # Configure the Text widget to use the scrollbars
        self.scrape_result_text.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set, wrap="none")
        
        # Place widgets in the frame
        self.scrape_result_text.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        # Make the Text widget expandable within its frame
        results_frame.rowconfigure(0, weight=1)
        results_frame.columnconfigure(0, weight=1)

    def browse_directory(self, entry_widget):
        directory = filedialog.askdirectory()
        if directory:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, directory)
    
    def browse_file(self, entry_widget, file_types):
        file_path = filedialog.askopenfilename(filetypes=file_types)
        if file_path:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, file_path)
    
    def check_unread_emails(self):
        try:
            # Clear previous results
            self.outlook_result_text.delete(1.0, tk.END)
            
            category = self.category_entry.get()
            if not category:
                messagebox.showerror("Error", "Please enter an email category.")
                return
                
            # Create OutlookProcessor instance
            processor = OutlookProcessor(
                category,
                [],  # Empty senders list for just checking
                "",  # Empty path for just checking
                ""   # Empty path for just checking
            )
            
            # Check unread emails count
            unread_count = processor.list_unread_emails()
            
            self.outlook_result_text.insert(tk.END, f"Found {unread_count} unread emails with category '{category}'.\n")
            self.status_var.set(f"Found {unread_count} unread emails")
            
        except Exception as e:
            self.outlook_result_text.insert(tk.END, f"Error: {str(e)}\n")
            self.status_var.set("Error checking emails")
    
    def process_emails(self):
        try:
            # Clear previous results
            self.outlook_result_text.delete(1.0, tk.END)
            
            category = self.category_entry.get()
            senders = [s.strip() for s in self.senders_entry.get().split(",") if s.strip()]
            attachment_path = self.attachment_path_entry.get()
            msg_path = self.msg_path_entry.get()
            
            if not category or not senders or not attachment_path or not msg_path:
                messagebox.showerror("Error", "Please fill in all required fields.")
                return
                
            self.outlook_result_text.insert(tk.END, "Starting email processing...\n")
            self.status_var.set("Processing emails...")
            self.root.update()
            
            # Create OutlookProcessor instance
            processor = OutlookProcessor(
                category,
                senders,
                attachment_path,
                msg_path
            )
            
            # Process emails in a separate thread to keep the UI responsive
            def process_thread():
                processor.download_attachments_and_save_as_msg(
                    self.save_emails_var.get(),
                    self.mark_as_read_var.get()
                )
                
                # Update UI from the main thread
                self.root.after(0, lambda: self.update_outlook_results(processor))
            
            threading.Thread(target=process_thread, daemon=True).start()
            
        except Exception as e:
            self.outlook_result_text.insert(tk.END, f"Error: {str(e)}\n")
            self.status_var.set("Error processing emails")
    
    def update_outlook_results(self, processor):
        self.outlook_result_text.insert(tk.END, "Email processing completed.\n\n")
        
        # Display results
        self.outlook_result_text.insert(tk.END, f"Emails processed: {len(processor.processed_emails)}\n")
        
        if processor.emails_with_pdf:
            self.outlook_result_text.insert(tk.END, "\nEmails with PDF attachments:\n")
            for subject in processor.emails_with_pdf:
                self.outlook_result_text.insert(tk.END, f"- {subject}\n")
                
        if processor.emails_with_nvf_new_vendor:
            self.outlook_result_text.insert(tk.END, "\nEmails with NVF or New Vendor attachments:\n")
            for subject in processor.emails_with_nvf_new_vendor:
                self.outlook_result_text.insert(tk.END, f"- {subject}\n")
        
        self.status_var.set("Email processing completed")
    
    def process_case_list(self):
        try:
            # Clear previous results
            self.case_list_result_text.delete(1.0, tk.END)
            
            excel_folder = self.excel_folder_entry.get()
            list_folder = self.list_folder_entry.get()
            
            if not excel_folder or not list_folder:
                messagebox.showerror("Error", "Please select both folders.")
                return
                
            self.case_list_result_text.insert(tk.END, "Processing case list...\n")
            self.status_var.set("Processing case list...")
            self.root.update()
            
            # Create CaseList instance
            case_list = CaseList(excel_folder, list_folder)
            
            # Process in a separate thread
            def process_thread():
                duplicate_counts, error_messages = case_list.process_excel_files()
                
                # Update UI from the main thread
                self.root.after(0, lambda: self.update_case_list_results(duplicate_counts, error_messages))
            
            threading.Thread(target=process_thread, daemon=True).start()
            
        except Exception as e:
            self.case_list_result_text.insert(tk.END, f"Error: {str(e)}\n")
            self.status_var.set("Error processing case list")
    
    def update_case_list_results(self, duplicate_counts, error_messages):
        self.case_list_result_text.insert(tk.END, "Case list processing completed.\n\n")
        
        # Show duplicates
        duplicates = sum(1 for count in duplicate_counts.values() if count > 0)
        self.case_list_result_text.insert(tk.END, f"Found {duplicates} duplicate cases.\n")
        
        if duplicates > 0:
            self.case_list_result_text.insert(tk.END, "\nDuplicate cases:\n")
            for value, count in duplicate_counts.items():
                if count > 0:
                    self.case_list_result_text.insert(tk.END, f"- {value} (Duplicated {count} times)\n")
        
        # Show errors
        if error_messages:
            self.case_list_result_text.insert(tk.END, "\nErrors encountered:\n")
            for error in error_messages:
                self.case_list_result_text.insert(tk.END, f"- {error}\n")
        
        self.status_var.set("Case list processing completed")
    
    def transpose_excel(self):
        try:
            # Clear previous results
            self.transpose_result_text.delete(1.0, tk.END)
            
            excel_file = self.excel_file_entry.get()
            sheet_name = self.sheet_name_entry.get().strip()
            
            if not excel_file:
                messagebox.showerror("Error", "Please select an Excel file.")
                return
                
            self.transpose_result_text.insert(tk.END, "Transposing Excel data...\n")
            self.status_var.set("Transposing Excel...")
            self.root.update()
            
            # Create ExcelTransposer instance
            transposer = ExcelTransposer(excel_file)
            
            # Set active sheet if specified
            if sheet_name:
                try:
                    transposer.set_active_sheet(sheet_name)
                except ValueError as e:
                    self.transpose_result_text.insert(tk.END, f"Error: {str(e)}\n")
                    self.status_var.set("Error transposing Excel")
                    return
            
            # Transpose in a separate thread
            def transpose_thread():
                transposer.transpose_cells_to_table()
                
                # Update UI from the main thread
                self.root.after(0, lambda: self.update_transpose_results(excel_file))
            
            threading.Thread(target=transpose_thread, daemon=True).start()
            
        except Exception as e:
            self.transpose_result_text.insert(tk.END, f"Error: {str(e)}\n")
            self.status_var.set("Error transposing Excel")
    
    def update_transpose_results(self, excel_file):
        self.transpose_result_text.insert(tk.END, "Excel transposition completed.\n\n")
        self.transpose_result_text.insert(tk.END, f"Transposed data saved to '{excel_file}' in a new sheet named 'Transposed'.\n")
        self.status_var.set("Excel transposition completed")
    
    def scrape_excel_files(self):
        try:
            # Clear previous results
            self.scrape_result_text.delete(1.0, tk.END)
            
            directory = self.scrape_dir_entry.get()
            range_start = self.range_start_entry.get()
            range_end = self.range_end_entry.get()
            read_headers = self.read_headers_var.get()
            
            if not directory:
                messagebox.showerror("Error", "Please select a directory with Excel files.")
                return
                
            self.scrape_result_text.insert(tk.END, f"Scraping Excel files in {directory}...\n")
            self.status_var.set("Scraping Excel files...")
            self.root.update()
            
            # Set directory for Excel scraper
            self.excel_scraper.set_directory(directory)
            
            # Scrape in a separate thread
            def scrape_thread():
                results = self.excel_scraper.scrape_excel_files(range_start, range_end, read_headers)
                
                # Update UI from the main thread
                self.root.after(0, lambda: self.update_scrape_results(results))
            
            threading.Thread(target=scrape_thread, daemon=True).start()
            
        except Exception as e:
            self.scrape_result_text.insert(tk.END, f"Error: {str(e)}\n")
            self.status_var.set("Error scraping Excel files")
    
    def update_scrape_results(self, results):
        self.scrape_result_text.delete(1.0, tk.END)
        self.scrape_result_text.insert(tk.END, f"Scraped {len(results)} Excel files.\n\n")
        
        # Display headers
        headers = self.excel_scraper.get_headers()
        if headers:
            self.scrape_result_text.insert(tk.END, "Headers found: " + ", ".join(headers) + "\n\n")
        
        # Display results in a tabular format
        if results:
            # Create header line
            header_line = f"{'Filename':<30} | "
            if headers:
                for header in headers:
                    header_line += f"{str(header):<15} | "
            else:
                # Use keys from the first result
                for key in results[0]["values"].keys():
                    header_line += f"{str(key):<15} | "
            
            self.scrape_result_text.insert(tk.END, header_line + "\n")
            self.scrape_result_text.insert(tk.END, "-" * len(header_line) + "\n")
            
            # Add each result row
            for result in results:
                line = f"{result['filename']:<30} | "
                for key, value in result["values"].items():
                    if value is None:
                        value = ""
                    line += f"{str(value):<15} | "
                self.scrape_result_text.insert(tk.END, line + "\n")
        
        self.status_var.set(f"Scraped {len(results)} Excel files")
    
    def export_to_csv(self):
        try:
            if not self.excel_scraper.get_results():
                messagebox.showerror("Error", "No data to export. Please scrape Excel files first.")
                return
                
            output_file = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
            )
            
            if not output_file:
                return  # User cancelled
                
            success = self.excel_scraper.save_results_to_csv(output_file)
            
            if success:
                messagebox.showinfo("Success", f"Data exported to {output_file}")
                self.status_var.set("Data exported to CSV")
            else:
                messagebox.showerror("Error", "Failed to export data")
                
        except Exception as e:
            messagebox.showerror("Error", f"Export error: {str(e)}")
            self.status_var.set("Error exporting data")

# Main application entry point
if __name__ == "__main__":
    try:
        root = Tk()
        app = ExcelProcessorApp(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("Application Error", f"An unexpected error occurred: {str(e)}")